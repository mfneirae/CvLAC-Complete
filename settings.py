#
#
# #############################################################################
#       Copyright (c) 2018 Universidad Nacional de Colombia All Rights Reserved.
#
#             This work was made as a development to improve data collection
#       for self-assessment and accreditation processes in the Vicedeanship
#       of academic affairs in the Engineering Faculty of the Universidad
#       Nacional de Colombia and is licensed under a Creative Commons
#       Attribution-NonCommercial - ShareAlike 4.0 International License
#       and MIT Licence.
#
#       by Manuel Embus.
#
#       For more information write me to jai@mfneirae.com
#       Or visit my webpage at https://mfneirae.com/
# #############################################################################
#
#

import openpyxl, sys, os, time, logging
start_time = time.time()
Dir = os.getcwd()
os.chdir(Dir+"/Bin")
sys.path.append('../Bin/')
import apropiacion
import init
import produccion_bibliografica
import produccion_tecnica
import comites
import reconocimientos
os.chdir(Dir)
global COD_PRODUCTO, COD_RECONOCIMIENTO
condition = 0;
while condition != 1:
    try:
        print ("------> Seleccione la forma en la que desea obtener la información:")
        print ("1) Imprimir datos en CSV y en Insert")
        print ("2) Imprimir datos en Insert para MySQL")
        print ("3) Imprimir datos en CSV")
        mode = int(input('-> Seleccione una opción: '))
        if mode == 1 or mode == 2 or mode == 3:
            condition = 1
        else:
            print ("El varlor escogido no es valido")
    except ValueError:
        print ("Not a number")
print ("------> Inicio de Importación de Registros.")
wb = openpyxl.load_workbook('./Input/Base.xlsx')
# wb = openpyxl.load_workbook('./Input/Base.xlsx')
sheet = wb['Sheet1']
total = sheet.max_row +1
COD_PRODUCTO = 1;
COD_RECONOCIMIENTO= 1;
init.inicio()
LOG_FILENAME = './Logs/Registros.log'
logging.basicConfig(filename=LOG_FILENAME,level=logging.DEBUG,
                        format = '%(asctime)s:%(levelname)s:%(message)s')
LEVELS = {'debug': logging.DEBUG,
          'info': logging.INFO,
          'warning': logging.WARNING,
          'error': logging.ERROR,
          'critical': logging.CRITICAL}
if len(sys.argv) > 1:
    level_name = sys.argv[1]
    level = LEVELS.get(level_name, logging.NOTSET)
    logging.basicConfig(level=level)
print("------> Los CvLAC han sido cargados, Estado: " + str(1/(total-1)*100) + "%")
for q in range(2,total):
    doc = sheet['A'+str(q)].value
    name = sheet['B'+str(q)].value
    last = sheet['C'+str(q)].value
    my_url = sheet['D'+str(q)].value
    index1 = my_url.find("cod_rh=") + 7
    index2 = len(my_url)
    RH = my_url[index1:index2]
    init.rel_persona_colciencias.append(str(doc) + ";"\
    + str(RH) \
    + "\n")
    init.inrel_persona_colciencias.append("REPLACE INTO `uapa_db`.`rel_persona_colciencias` (`cod_rh`,`dni_persona`) VALUES " \
    + "('" + str(RH) + "','" + str(doc) + "');\n")
    if my_url != '-':
        apropiacion.evenextract()
        from apropiacion import conteventos
        COD_PRODUCTO = int("".join(str(x) for x in conteventos))
        apropiacion.redesextract()
        from apropiacion import contredes
        COD_PRODUCTO = int("".join(str(x) for x in contredes))
        apropiacion.estrategiaextract()
        from apropiacion import contEstrategia
        COD_PRODUCTO = int("".join(str(x) for x in contEstrategia))
        produccion_bibliografica.artiextract()
        from produccion_bibliografica import contarticulo
        COD_PRODUCTO = int("".join(str(x) for x in contarticulo))
        produccion_bibliografica.libextract()
        from produccion_bibliografica import contlibro
        COD_PRODUCTO = int("".join(str(x) for x in contlibro))
        produccion_bibliografica.caplibroextract()
        from produccion_bibliografica import contcaplibro
        COD_PRODUCTO = int("".join(str(x) for x in contcaplibro))
        produccion_bibliografica.texnocienextract()
        from produccion_bibliografica import conttexnocien
        COD_PRODUCTO = int("".join(str(x) for x in conttexnocien))
        produccion_bibliografica.workpextract()
        from produccion_bibliografica import contworkp
        COD_PRODUCTO = int("".join(str(x) for x in contworkp))
        produccion_bibliografica.traduccionextract()
        from produccion_bibliografica import conttraduccion
        COD_PRODUCTO = int("".join(str(x) for x in conttraduccion))
        produccion_bibliografica.otrapbextract()
        from produccion_bibliografica import contotrapb
        COD_PRODUCTO = int("".join(str(x) for x in contotrapb))
        produccion_tecnica.ptsoftwareextract()
        from produccion_tecnica import contptsoftware
        COD_PRODUCTO = int("".join(str(x) for x in contptsoftware))
        produccion_tecnica.ptproductoextract()
        from produccion_tecnica import contptproducto
        COD_PRODUCTO = int("".join(str(x) for x in contptproducto))
        produccion_tecnica.ptdiseñoextract()
        from produccion_tecnica import contptdiseño
        COD_PRODUCTO = int("".join(str(x) for x in contptdiseño))
        produccion_tecnica.ptcircuitosextract()
        from produccion_tecnica import contptcircuitos
        COD_PRODUCTO = int("".join(str(x) for x in contptcircuitos))
        produccion_tecnica.ptinnovaextract()
        from produccion_tecnica import contptinnova
        COD_PRODUCTO = int("".join(str(x) for x in contptinnova))
        produccion_tecnica.ptvaranimalextract()
        from produccion_tecnica import contptvaranimal
        COD_PRODUCTO = int("".join(str(x) for x in contptvaranimal))
        produccion_tecnica.ptprocesoextract()
        from produccion_tecnica import contptproceso
        COD_PRODUCTO = int("".join(str(x) for x in contptproceso))
        produccion_tecnica.ptcartasextract()
        from produccion_tecnica import contptcartas
        COD_PRODUCTO = int("".join(str(x) for x in contptcartas))
        produccion_tecnica.ptvegetalextract()
        from produccion_tecnica import contptvegetal
        COD_PRODUCTO = int("".join(str(x) for x in contptvegetal))
        produccion_tecnica.pttratecextract()
        from produccion_tecnica import contpttratec
        COD_PRODUCTO = int("".join(str(x) for x in contpttratec))
        produccion_tecnica.ptnormaextract()
        from produccion_tecnica import contptnorma
        COD_PRODUCTO = int("".join(str(x) for x in contptnorma))
        produccion_tecnica.ptreglamentoextract()
        from produccion_tecnica import contptreglamento
        COD_PRODUCTO = int("".join(str(x) for x in contptreglamento))
        produccion_tecnica.ptempresaextract()
        from produccion_tecnica import contptempresa
        COD_PRODUCTO = int("".join(str(x) for x in contptempresa))
        produccion_tecnica.demastrabajosextract()
        from produccion_tecnica import contdemastrabajos
        COD_PRODUCTO = int("".join(str(x) for x in contdemastrabajos))
        produccion_tecnica.ptsignosextract()
        from produccion_tecnica import contptsignos
        COD_PRODUCTO = int("".join(str(x) for x in contptsignos))
        comites.pcomitesextract()
        from comites import contpcomites
        COD_PRODUCTO = int("".join(str(x) for x in contpcomites))
        comites.pjcomitesextract()
        from comites import contpjcomites
        COD_PRODUCTO = int("".join(str(x) for x in contpjcomites))
        reconocimientos.reconocimientosextract()
        from reconocimientos import contreconocimientos
        COD_PRODUCTO = int("".join(str(x) for x in contreconocimientos))
        print("------> "+ name + " " + last + " ha sido procesado, Estado: " + str(q/(total-1)*100) + "%")
        if q==total-1:
            logging.shutdown()
            print ("------> Escribiendo las bases de datos.")
            if mode == 1:
                import printcsv
                import printinsert
            elif mode == 2:
                import printinsert
            else:
                import printcsv
            print ("-----------------------------------------------------------------------------------------------")
            print ("")
            print ("------> ¡Extracción Exitosa!")
            print ("------> La información se encuentra en la carpeta: Resultados.")
            print ("------> Tiempo de ejecución: %s Minutos." % ((time.time() - start_time)/60))
            print ("")
            print ("***********************************************************************************************")
            sys.exit()
    COD_PRODUCTO = 1;
    COD_RECONOCIMIENTO = 1;
